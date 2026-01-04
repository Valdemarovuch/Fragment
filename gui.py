"""
Графічний інтерфейс для Fragment Corruptor
"""
import customtkinter as ctk
import tkinter as tk
from tkinter import filedialog, messagebox
import os
import shutil
import threading
import logging
from datetime import datetime
import subprocess
import stat
import tempfile
import zipfile

from config import CorruptionMode, AppConfig, NamingPattern, OperationMode, DeletionMode
from corruption_engine import CorruptionEngine
from secure_delete import SecureShredder

# Імпорт бібліотек для Office документів
try:
    from docx import Document
    DOCX_AVAILABLE = True
except ImportError:
    DOCX_AVAILABLE = False

try:
    from openpyxl import load_workbook
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False

try:
    from pptx import Presentation
    PPTX_AVAILABLE = True
except ImportError:
    PPTX_AVAILABLE = False

try:
    from PIL import Image
    PIL_AVAILABLE = True
except ImportError:
    PIL_AVAILABLE = False


class FileBreakerApp(ctk.CTk):
    """Основний клас додатку"""
    
    def __init__(self):
        super().__init__()
        
        # Налаштування вікна
        self.title(AppConfig.WINDOW_TITLE)
        self.geometry(f"{AppConfig.WINDOW_WIDTH}x{AppConfig.WINDOW_HEIGHT}")
        self.minsize(AppConfig.WINDOW_MIN_WIDTH, AppConfig.WINDOW_MIN_HEIGHT)
        
        # Дані додатку
        self.selected_files = []
        self.file_rows = {}
        
        # Параметри операції
        self.operation_mode = tk.StringVar(value=OperationMode.CORRUPT)
        
        # Параметри корупції
        self.corruption_mode = tk.StringVar(value=CorruptionMode.FULL)
        self.output_directory = tk.StringVar(value="")
        self.custom_header_bytes = tk.IntVar(value=16)
        self.custom_noise_count = tk.IntVar(value=5)
        self.custom_truncate_bytes = tk.IntVar(value=30)
        
        # Параметри безпечного видалення
        self.deletion_mode = tk.StringVar(value=DeletionMode.STANDARD)
        self.deletion_passes = tk.IntVar(value=3)
        
        # Налаштування імен файлів
        self.naming_pattern = tk.StringVar(value=NamingPattern.SAME)
        self.custom_prefix = tk.StringVar(value="")
        self.custom_suffix = tk.StringVar(value="")
        
        # Статистика
        self.total_stats = {
            'total_files': 0,
            'total_bytes_changed': 0,
            'total_time': 0.0,
            'original_total_size': 0,
            'corrupted_total_size': 0
        }
        
        logging.info("Application started")
        
        # Побудова інтерфейсу
        self._build_ui()
    
    def _build_ui(self):
        """Побудова інтерфейсу користувача"""
        self._create_header()
        self._create_controls()
        self._create_file_list()
        self._create_footer()
    
    def _create_header(self):
        """Створення заголовку"""
        header_frame = ctk.CTkFrame(self, height=60, corner_radius=0, fg_color="transparent")
        header_frame.pack(fill="x", padx=20, pady=(20, 0))
        
        logo_label = ctk.CTkLabel(header_frame, text=AppConfig.HEADER_TEXT, 
                                 font=("Roboto Medium", 20))
        logo_label.pack(side="top")
    
    def _create_controls(self):
        """Створення панелі керування"""
        controls_frame = ctk.CTkFrame(self, height=50, corner_radius=0, fg_color="transparent")
        controls_frame.pack(fill="x", padx=20, pady=10)
        
        # Кнопки додавання
        btn_add = ctk.CTkButton(controls_frame, text="+ Add Files", width=120, height=35,
                               fg_color="white", text_color="black", hover_color="#e0e0e0",
                               command=self.add_files)
        btn_add.pack(side="left")
        
        btn_add_folder = ctk.CTkButton(controls_frame, text="+ Add Folder", width=120, height=35,
                                      fg_color="white", text_color="black", hover_color="#e0e0e0",
                                      command=self.add_folder)
        btn_add_folder.pack(side="left", padx=(10, 0))
        
        # Статус
        self.lbl_status = ctk.CTkLabel(controls_frame, text="0 files loaded", text_color="gray")
        self.lbl_status.pack(side="left", padx=20)
        
        # Кнопка очищення
        btn_clear = ctk.CTkButton(controls_frame, text="🗑 Clear All", width=80,
                                 fg_color="transparent", border_width=1, border_color="gray",
                                 text_color="gray", hover_color="#333333",
                                 command=self.clear_list)
        btn_clear.pack(side="right")
    
    def _create_file_list(self):
        """Створення списку файлів"""
        self.file_list_frame = ctk.CTkScrollableFrame(self, fg_color="#1a1a1a", corner_radius=10)
        self.file_list_frame.pack(fill="both", expand=True, padx=20, pady=10)
        
        # Заголовки таблиці
        self.list_header = ctk.CTkFrame(self.file_list_frame, height=30, fg_color="transparent")
        self.list_header.pack(fill="x", pady=(0, 5))
        
        ctk.CTkLabel(self.list_header, text="FILENAME", font=("Roboto", 12, "bold"),
                    text_color="gray", anchor="w").pack(side="left", padx=10, fill="x", expand=True)
        ctk.CTkLabel(self.list_header, text="SIZE", font=("Roboto", 12, "bold"),
                    text_color="gray", width=80).pack(side="left", padx=5)
        ctk.CTkLabel(self.list_header, text="STATUS", font=("Roboto", 12, "bold"),
                    text_color="gray", width=100).pack(side="left", padx=5)
        ctk.CTkLabel(self.list_header, text="", width=30).pack(side="right", padx=5)
    
    def _create_footer(self):
        """Створення підвалу з налаштуваннями"""
        footer_frame = ctk.CTkFrame(self, height=200, corner_radius=0, fg_color="#111111")
        footer_frame.pack(fill="x", side="bottom")
        
        # Прогрес-бар
        self.progress_bar = ctk.CTkProgressBar(footer_frame, height=3)
        self.progress_bar.pack(fill="x", side="top")
        self.progress_bar.set(0)
        
        footer_inner = ctk.CTkFrame(footer_frame, fg_color="transparent")
        footer_inner.pack(fill="both", expand=True, padx=20, pady=15)
        
        # Налаштування (ліва частина)
        settings_frame = ctk.CTkFrame(footer_inner, fg_color="transparent")
        settings_frame.pack(side="left", fill="both", expand=True)
        
        # ПЕРШИЙ РЯДОК - Вибір операції (Corrupt vs Secure Delete)
        operation_frame = ctk.CTkFrame(settings_frame, fg_color="transparent")
        operation_frame.pack(anchor="w", fill="x", pady=(0, 10))
        
        ctk.CTkLabel(operation_frame, text="Operation:", text_color="gray",
                    font=("Roboto", 11)).pack(side="left")
        
        self.operation_menu = ctk.CTkOptionMenu(
            operation_frame,
            variable=self.operation_mode,
            values=[OperationMode.CORRUPT, OperationMode.SECURE_DELETE],
            width=200,
            command=self.on_operation_change
        )
        self.operation_menu.pack(side="left", padx=(10, 20))
        
        # ДРУГИЙ РЯДОК - режим corruption (видимо лише для CORRUPT)
        self.corruption_frame = ctk.CTkFrame(settings_frame, fg_color="transparent")
        self.corruption_frame.pack(anchor="w", fill="x", pady=(0, 10))
        
        ctk.CTkLabel(self.corruption_frame, text="Corruption Mode:", text_color="gray",
                    font=("Roboto", 11)).pack(anchor="w")
        
        self.corruption_mode_menu = ctk.CTkOptionMenu(
            self.corruption_frame,
            variable=self.corruption_mode,
            values=[CorruptionMode.FULL, CorruptionMode.HEADER_ONLY, 
                   CorruptionMode.NOISE_ONLY, CorruptionMode.TRUNCATE_ONLY, 
                   CorruptionMode.CUSTOM],
            width=300,
            command=self.on_mode_change
        )
        self.corruption_mode_menu.pack(anchor="w", pady=(5, 0))
        
        # ТРЕТІЙ РЯДОК - режим безпечного видалення (видимо лише для SECURE_DELETE)
        self.deletion_frame = ctk.CTkFrame(settings_frame, fg_color="transparent")
        self.deletion_frame.pack(anchor="w", fill="x", pady=(0, 10))
        
        deletion_row1 = ctk.CTkFrame(self.deletion_frame, fg_color="transparent")
        deletion_row1.pack(anchor="w", fill="x")
        
        ctk.CTkLabel(deletion_row1, text="Deletion Mode:", text_color="gray",
                    font=("Roboto", 11)).pack(side="left")
        
        self.deletion_mode_menu = ctk.CTkOptionMenu(
            deletion_row1,
            variable=self.deletion_mode,
            values=[DeletionMode.QUICK, DeletionMode.STANDARD, DeletionMode.EXTENDED],
            width=200,
            command=self.on_deletion_mode_change
        )
        self.deletion_mode_menu.pack(side="left", padx=(10, 20))
        
        # Контроль для кількості проходів
        ctk.CTkLabel(deletion_row1, text="Passes:", text_color="gray",
                    font=("Roboto", 11)).pack(side="left", padx=(10, 0))
        
        self.deletion_passes_spinbox = ctk.CTkSpinbox(
            deletion_row1,
            from_=1, to=35,
            variable=self.deletion_passes,
            width=60
        )
        self.deletion_passes_spinbox.pack(side="left", padx=(5, 0))
        
        # Приховуємо deletion frame за замовчуванням
        self.deletion_frame.pack_forget()
        
        # ЧЕТВЕРТИЙ РЯДОК - naming pattern
        naming_frame = ctk.CTkFrame(settings_frame, fg_color="transparent")
        naming_frame.pack(fill="x", pady=(0, 10))
        
        ctk.CTkLabel(naming_frame, text="File Naming:", text_color="gray",
                    font=("Roboto", 11)).pack(side="left")
        
        self.naming_menu = ctk.CTkOptionMenu(
            naming_frame,
            variable=self.naming_pattern,
            values=[NamingPattern.SAME, NamingPattern.NUMBERED, NamingPattern.PREFIX,
                   NamingPattern.SUFFIX, NamingPattern.TIMESTAMP, NamingPattern.CUSTOM],
            width=180,
            command=self.on_naming_change
        )
        self.naming_menu.pack(side="left", padx=(10, 5))
        
        # Поля для кастомних prefix/suffix
        self.prefix_entry = ctk.CTkEntry(naming_frame, placeholder_text="Prefix", width=80)
        self.suffix_entry = ctk.CTkEntry(naming_frame, placeholder_text="Suffix", width=80)
        
        # П'ЯТИ РЯДОК - вибір output папки (лише для CORRUPT)
        self.output_frame = ctk.CTkFrame(settings_frame, fg_color="transparent")
        self.output_frame.pack(fill="x")
        
        ctk.CTkLabel(self.output_frame, text="Output:", text_color="gray",
                    font=("Roboto", 11)).pack(side="left")
        self.lbl_output = ctk.CTkLabel(self.output_frame, text="Same as original",
                                      text_color="#00aaff", font=("Roboto", 10))
        self.lbl_output.pack(side="left", padx=(5, 10))
        
        btn_output = ctk.CTkButton(self.output_frame, text="Change", width=80, height=25,
                                  fg_color="transparent", border_width=1, border_color="gray",
                                  command=self.select_output_directory)
        btn_output.pack(side="left")
        
        # Кнопка запуску (права частина)
        self.btn_run = ctk.CTkButton(footer_inner, text="Corrupt All →", width=150, height=40,
                                     font=("Roboto Medium", 14),
                                     command=self.start_corruption_thread)
        self.btn_run.pack(side="right", anchor="e")
    
    # --- МЕТОДИ КЕРУВАННЯ ФАЙЛАМИ ---
    
    def add_files(self):
        """Додавання файлів через діалог"""
        files = filedialog.askopenfilenames(
            title="Select files to corrupt",
            filetypes=[("All Files", "*.*"), ("Word Documents", "*.docx"),
                      ("Images", "*.jpg *.png"), ("PDFs", "*.pdf")]
        )
        if files:
            for file_path in files:
                self.add_file_to_list(file_path)
    
    def add_folder(self):
        """Додавання всіх файлів з папки"""
        folder = filedialog.askdirectory(title="Select folder")
        if folder:
            added_count = 0
            for root, dirs, files in os.walk(folder):
                for file in files:
                    if not file.startswith('.'):
                        file_path = os.path.join(root, file)
                        if self.add_file_to_list(file_path):
                            added_count += 1
            logging.info(f"Added {added_count} files from folder: {folder}")
    
    def add_file_to_list(self, file_path):
        """Додавання одного файлу до списку"""
        if not os.path.exists(file_path):
            logging.warning(f"File does not exist: {file_path}")
            return False
        
        if file_path in self.selected_files:
            logging.debug(f"File already in list: {file_path}")
            return False
        
        if not os.access(file_path, os.R_OK):
            logging.warning(f"Cannot read file: {file_path}")
            messagebox.showwarning("Warning", f"Cannot read file:\n{os.path.basename(file_path)}")
            return False
        
        self.selected_files.append(file_path)
        self.create_file_row(file_path)
        self.update_status_label()
        return True
    
    def create_file_row(self, file_path):
        """Створення рядка файлу в списку"""
        filename = os.path.basename(file_path)
        file_size = os.path.getsize(file_path)
        size_str = self.format_file_size(file_size)
        
        row_frame = ctk.CTkFrame(self.file_list_frame, height=40, fg_color="#2b2b2b", corner_radius=5)
        row_frame.pack(fill="x", pady=2)
        
        lbl_name = ctk.CTkLabel(row_frame, text=filename, anchor="w", padx=10)
        lbl_name.pack(side="left", fill="x", expand=True)
        
        lbl_size = ctk.CTkLabel(row_frame, text=size_str, text_color="gray", width=80)
        lbl_size.pack(side="left", padx=5)
        
        lbl_status = ctk.CTkLabel(row_frame, text="Ready", text_color=AppConfig.COLOR_READY, width=100)
        lbl_status.pack(side="left", padx=5)
        
        btn_remove = ctk.CTkButton(row_frame, text="×", width=30, height=30,
                                  fg_color="transparent", hover_color="#ff5555",
                                  command=lambda: self.remove_file(file_path))
        btn_remove.pack(side="right", padx=5)
        
        row_frame.status_label = lbl_status
        row_frame.file_path = file_path
        self.file_rows[file_path] = row_frame
    
    def remove_file(self, file_path):
        """Видалення файлу зі списку"""
        if file_path in self.selected_files:
            self.selected_files.remove(file_path)
        
        if file_path in self.file_rows:
            self.file_rows[file_path].destroy()
            del self.file_rows[file_path]
        
        self.update_status_label()
        logging.info(f"Removed file: {file_path}")
    
    def clear_list(self):
        """Очищення списку файлів"""
        self.selected_files = []
        self.file_rows = {}
        for widget in self.file_list_frame.winfo_children():
            if isinstance(widget, ctk.CTkFrame) and widget != self.list_header:
                widget.destroy()
        self.update_status_label()
        self.progress_bar.set(0)
        logging.info("Cleared all files")
    
    def update_status_label(self):
        """Оновлення лічильника файлів"""
        count = len(self.selected_files)
        self.lbl_status.configure(text=f"{count} file{'s' if count != 1 else ''} loaded")
    
    @staticmethod
    def format_file_size(size_bytes):
        """Форматування розміру файлу"""
        for unit in ['B', 'KB', 'MB', 'GB']:
            if size_bytes < 1024.0:
                return f"{size_bytes:.1f} {unit}"
            size_bytes /= 1024.0
        return f"{size_bytes:.1f} TB"
    
    def select_output_directory(self):
        """Вибір вихідної директорії"""
        directory = filedialog.askdirectory(title="Select output directory")
        if directory:
            self.output_directory.set(directory)
            self.lbl_output.configure(text=os.path.basename(directory))
            logging.info(f"Output directory set to: {directory}")
        else:
            self.output_directory.set("")
            self.lbl_output.configure(text="Same as original")
    
    def on_mode_change(self, choice):
        """Обробка зміни режиму corruption"""
        logging.info(f"Corruption mode changed to: {choice}")
    
    def on_deletion_mode_change(self, choice):
        """Обробка зміни режиму безпечного видалення"""
        # Автоматично встановлюємо кількість проходів на основі режиму
        if choice == DeletionMode.QUICK:
            self.deletion_passes.set(1)
        elif choice == DeletionMode.STANDARD:
            self.deletion_passes.set(3)
        elif choice == DeletionMode.EXTENDED:
            self.deletion_passes.set(7)
        logging.info(f"Deletion mode changed to: {choice} ({self.deletion_passes.get()} passes)")
    
    def on_operation_change(self, choice):
        """Обробка зміни режиму операції (Corrupt vs Secure Delete)"""
        if choice == OperationMode.CORRUPT:
            # Показуємо controls для corruption
            self.corruption_frame.pack(anchor="w", fill="x", pady=(0, 10), after=self.operation_menu.master)
            self.output_frame.pack(fill="x", after=self.naming_menu.master)
            self.naming_menu.pack()
            
            # Приховуємо controls для deletion
            self.deletion_frame.pack_forget()
            
            # Оновлюємо кнопку запуску
            self.btn_run.configure(text="Corrupt All →")
        
        elif choice == OperationMode.SECURE_DELETE:
            # Приховуємо controls для corruption
            self.corruption_frame.pack_forget()
            self.output_frame.pack_forget()
            
            # Показуємо controls для deletion
            self.deletion_frame.pack(anchor="w", fill="x", pady=(0, 10), 
                                     after=self.operation_menu.master)
            
            # Приховуємо naming pattern для delete режиму
            self.naming_menu.pack_forget()
            
            # Оновлюємо кнопку запуску
            self.btn_run.configure(text="Secure Delete →")
        
        logging.info(f"Operation mode changed to: {choice}")
    
    def on_naming_change(self, choice):
        """Обробка зміни шаблону імен"""
        # Показуємо/приховуємо поля для custom
        if choice == NamingPattern.CUSTOM:
            self.prefix_entry.pack(side="left", padx=2)
            self.suffix_entry.pack(side="left", padx=2)
        else:
            self.prefix_entry.pack_forget()
            self.suffix_entry.pack_forget()
        logging.info(f"Naming pattern changed to: {choice}")
    
    def generate_output_filename(self, original_path, directory):
        """
        Генерація імені вихідного файлу на основі обраного шаблону
        
        Args:
            original_path: Шлях до оригінального файлу
            directory: Директорія для збереження
            
        Returns:
            str: Повний шлях до нового файлу
        """
        filename = os.path.basename(original_path)
        name, ext = os.path.splitext(filename)
        pattern = self.naming_pattern.get()
        
        if pattern == NamingPattern.SAME:
            new_filename = filename
        elif pattern == NamingPattern.NUMBERED:
            new_filename = filename
        elif pattern == NamingPattern.PREFIX:
            new_filename = f"[BROKEN] {filename}"
        elif pattern == NamingPattern.SUFFIX:
            new_filename = f"{name} [BROKEN]{ext}"
        elif pattern == NamingPattern.TIMESTAMP:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            new_filename = f"{name}_{timestamp}{ext}"
        elif pattern == NamingPattern.CUSTOM:
            prefix = self.prefix_entry.get()
            suffix = self.suffix_entry.get()
            new_filename = f"{prefix}{name}{suffix}{ext}"
        else:
            new_filename = filename
        
        new_path = os.path.join(directory, new_filename)
        
        # Якщо файл існує і pattern не TIMESTAMP, додаємо номер
        if os.path.exists(new_path) and pattern != NamingPattern.TIMESTAMP:
            counter = 1
            while os.path.exists(new_path):
                if pattern == NamingPattern.CUSTOM:
                    prefix = self.prefix_entry.get()
                    suffix = self.suffix_entry.get()
                    new_filename = f"{prefix}{name} ({counter}){suffix}{ext}"
                else:
                    new_filename = f"{name} ({counter}){ext}"
                new_path = os.path.join(directory, new_filename)
                counter += 1
        
        return new_path
    
    def preserve_metadata(self, source_path, target_path):
        """
        Копіює ВСІ метадані з оригіналу на цільовий файл
        
        Args:
            source_path: Шлях до оригінального файлу
            target_path: Шлях до цільового файлу
        """
        try:
            # Отримуємо розширення файлу
            ext = os.path.splitext(target_path)[1].lower()
            
            # Для Office документів використовуємо ТІЛЬКИ XML копіювання
            # щоб уникнути автоматичного оновлення дат при збереженні
            if ext in ['.docx', '.xlsx', '.pptx', '.docm', '.xlsm', '.pptm']:
                self._copy_office_metadata_xml(source_path, target_path)
            elif ext in ['.jpg', '.jpeg', '.png', '.gif', '.bmp', '.tiff'] and PIL_AVAILABLE:
                self._copy_image_metadata(source_path, target_path)
            
            # Системні метадані копіюємо В ОСТАННЮ ЧЕРГУ
            # щоб вони не перезаписались після маніпуляцій з файлом
            self._copy_system_metadata(source_path, target_path)
            
            logging.debug(f"Full metadata preserved for {target_path}")
            
        except Exception as e:
            logging.warning(f"Failed to preserve some metadata: {e}")
    
    def _copy_system_metadata(self, source_path, target_path):
        """
        Копіює системні метадані файлу (час створення, модифікації, права доступу)
        
        Args:
            source_path: Шлях до оригінального файлу
            target_path: Шлях до цільового файлу
        """
        try:
            source_stat = os.stat(source_path)
            
            # Копіюємо права доступу
            os.chmod(target_path, stat.S_IMODE(source_stat.st_mode))
            
            # Копіюємо modification time та access time з наносекундною точністю
            os.utime(target_path, ns=(source_stat.st_atime_ns, source_stat.st_mtime_ns))
            
            # Для Windows - копіюємо також creation time
            try:
                import win32file
                import win32con
                import pywintypes
                
                # Отримуємо час створення оригіналу
                handle = win32file.CreateFile(
                    source_path,
                    win32con.GENERIC_READ,
                    win32con.FILE_SHARE_READ,
                    None,
                    win32con.OPEN_EXISTING,
                    0,
                    None
                )
                creation_time, _, _ = win32file.GetFileTime(handle)
                win32file.CloseHandle(handle)
                
                # Встановлюємо час створення для цільового файлу
                handle = win32file.CreateFile(
                    target_path,
                    win32con.GENERIC_WRITE,
                    0,
                    None,
                    win32con.OPEN_EXISTING,
                    0,
                    None
                )
                win32file.SetFileTime(handle, creation_time, None, None)
                win32file.CloseHandle(handle)
                logging.debug(f"Windows creation time copied for {target_path}")
            except ImportError:
                # pywin32 не встановлено, пропускаємо
                pass
            except Exception as e:
                logging.debug(f"Could not set Windows creation time: {e}")
            
        except Exception as e:
            logging.warning(f"Failed to copy system metadata: {e}")
    
    def _copy_office_metadata_xml(self, source_path, target_path):
        """
        Копіює метадані Office документа через пряме копіювання XML файлів з архіву
        ПОВНІСТЮ 1:1 без змін дат
        
        Args:
            source_path: Шлях до оригінального файлу
            target_path: Шлях до цільового файлу
        """
        ext = os.path.splitext(target_path)[1].lower()
        
        # Перевіряємо чи це Office файл (zip архів)
        if ext not in ['.docx', '.xlsx', '.pptx', '.docm', '.xlsm', '.pptm']:
            return
        
        try:
            # Створюємо тимчасову директорію
            with tempfile.TemporaryDirectory() as temp_dir:
                # Розпаковуємо обидва файли
                source_extract = os.path.join(temp_dir, 'source')
                target_extract = os.path.join(temp_dir, 'target')
                os.makedirs(source_extract)
                os.makedirs(target_extract)
                
                # Розпаковуємо з збереженням часових міток
                with zipfile.ZipFile(source_path, 'r') as zip_ref:
                    zip_ref.extractall(source_extract)
                
                with zipfile.ZipFile(target_path, 'r') as zip_ref:
                    zip_ref.extractall(target_extract)
                
                # Копіюємо ВСІ файли з docProps (всі метадані)
                source_docprops = os.path.join(source_extract, 'docProps')
                target_docprops = os.path.join(target_extract, 'docProps')
                
                if os.path.exists(source_docprops):
                    # Видаляємо стару директорію з метаданими
                    if os.path.exists(target_docprops):
                        shutil.rmtree(target_docprops)
                    
                    # Копіюємо всю директорію docProps з оригіналу
                    shutil.copytree(source_docprops, target_docprops)
                    logging.debug(f"Copied entire docProps folder with all metadata")
                
                # Перепаковуємо цільовий файл з оновленими метаданими
                # Зберігаємо оригінальні часові мітки всередині архіву
                os.remove(target_path)
                with zipfile.ZipFile(target_path, 'w', zipfile.ZIP_DEFLATED) as zip_ref:
                    for root, dirs, files in os.walk(target_extract):
                        for file in files:
                            file_path = os.path.join(root, file)
                            arcname = os.path.relpath(file_path, target_extract)
                            
                            # Читаємо файл
                            with open(file_path, 'rb') as f:
                                file_data = f.read()
                            
                            # Отримуємо оригінальні часові мітки
                            file_stat = os.stat(file_path)
                            date_time = datetime.fromtimestamp(file_stat.st_mtime)
                            
                            # Створюємо ZipInfo з оригінальною датою
                            zinfo = zipfile.ZipInfo(
                                filename=arcname,
                                date_time=(date_time.year, date_time.month, date_time.day,
                                          date_time.hour, date_time.minute, date_time.second)
                            )
                            zinfo.compress_type = zipfile.ZIP_DEFLATED
                            
                            # Записуємо файл з оригінальною датою
                            zip_ref.writestr(zinfo, file_data)
                
                logging.debug(f"Office metadata XML copied 1:1 for {target_path}")
                
        except Exception as e:
            logging.debug(f"Could not copy Office metadata XML: {e}")
            import traceback
            logging.debug(traceback.format_exc())
    
    def _copy_docx_metadata(self, source_path, target_path):
        """
        Копіює метадані Word документа (.docx)
        
        Args:
            source_path: Шлях до оригінального файлу
            target_path: Шлях до цільового файлу
        """
        try:
            from docx import Document
            from docx.opc.coreprops import CoreProperties
            
            # Відкриваємо обидва документи
            source_doc = Document(source_path)
            target_doc = Document(target_path)
            
            # Копіюємо core properties
            source_props = source_doc.core_properties
            target_props = target_doc.core_properties
            
            # Копіюємо всі можливі властивості
            if hasattr(source_props, 'author') and source_props.author:
                target_props.author = source_props.author
            if hasattr(source_props, 'category') and source_props.category:
                target_props.category = source_props.category
            if hasattr(source_props, 'comments') and source_props.comments:
                target_props.comments = source_props.comments
            if hasattr(source_props, 'content_status') and source_props.content_status:
                target_props.content_status = source_props.content_status
            if hasattr(source_props, 'created') and source_props.created:
                target_props.created = source_props.created
            if hasattr(source_props, 'identifier') and source_props.identifier:
                target_props.identifier = source_props.identifier
            if hasattr(source_props, 'keywords') and source_props.keywords:
                target_props.keywords = source_props.keywords
            if hasattr(source_props, 'language') and source_props.language:
                target_props.language = source_props.language
            if hasattr(source_props, 'last_modified_by') and source_props.last_modified_by:
                target_props.last_modified_by = source_props.last_modified_by
            if hasattr(source_props, 'last_printed') and source_props.last_printed:
                target_props.last_printed = source_props.last_printed
            if hasattr(source_props, 'modified') and source_props.modified:
                target_props.modified = source_props.modified
            if hasattr(source_props, 'revision') and source_props.revision:
                target_props.revision = source_props.revision
            if hasattr(source_props, 'subject') and source_props.subject:
                target_props.subject = source_props.subject
            if hasattr(source_props, 'title') and source_props.title:
                target_props.title = source_props.title
            if hasattr(source_props, 'version') and source_props.version:
                target_props.version = source_props.version
            
            # Зберігаємо документ з оновленими метаданими
            target_doc.save(target_path)
            logging.debug(f"DOCX metadata copied for {target_path}")
            
        except Exception as e:
            logging.debug(f"Could not copy DOCX metadata using python-docx: {e}")
            # Fallback до копіювання XML
            self._copy_office_metadata_xml(source_path, target_path)
    
    def _copy_xlsx_metadata(self, source_path, target_path):
        """
        Копіює метадані Excel документа (.xlsx)
        
        Args:
            source_path: Шлях до оригінального файлу
            target_path: Шлях до цільового файлу
        """
        try:
            from openpyxl import load_workbook
            
            # Відкриваємо обидва документи
            source_wb = load_workbook(source_path)
            target_wb = load_workbook(target_path)
            
            # Копіюємо properties
            if hasattr(source_wb, 'properties'):
                source_props = source_wb.properties
                target_props = target_wb.properties
                
                if source_props.creator:
                    target_props.creator = source_props.creator
                if source_props.lastModifiedBy:
                    target_props.lastModifiedBy = source_props.lastModifiedBy
                if source_props.created:
                    target_props.created = source_props.created
                if source_props.modified:
                    target_props.modified = source_props.modified
                if source_props.title:
                    target_props.title = source_props.title
                if source_props.description:
                    target_props.description = source_props.description
                if source_props.subject:
                    target_props.subject = source_props.subject
                if source_props.keywords:
                    target_props.keywords = source_props.keywords
                if source_props.category:
                    target_props.category = source_props.category
                if source_props.contentStatus:
                    target_props.contentStatus = source_props.contentStatus
                if source_props.version:
                    target_props.version = source_props.version
                if source_props.revision:
                    target_props.revision = source_props.revision
                if source_props.lastPrinted:
                    target_props.lastPrinted = source_props.lastPrinted
            
            # Зберігаємо workbook
            target_wb.save(target_path)
            logging.debug(f"XLSX metadata copied for {target_path}")
            
        except Exception as e:
            logging.debug(f"Could not copy XLSX metadata using openpyxl: {e}")
            # Fallback до копіювання XML
            self._copy_office_metadata_xml(source_path, target_path)
    
    def _copy_pptx_metadata(self, source_path, target_path):
        """
        Копіює метадані PowerPoint документа (.pptx)
        
        Args:
            source_path: Шлях до оригінального файлу
            target_path: Шлях до цільового файлу
        """
        try:
            from pptx import Presentation
            
            # Відкриваємо обидві презентації
            source_prs = Presentation(source_path)
            target_prs = Presentation(target_path)
            
            # Копіюємо core properties
            source_props = source_prs.core_properties
            target_props = target_prs.core_properties
            
            if source_props.author:
                target_props.author = source_props.author
            if source_props.category:
                target_props.category = source_props.category
            if source_props.comments:
                target_props.comments = source_props.comments
            if source_props.content_status:
                target_props.content_status = source_props.content_status
            if source_props.created:
                target_props.created = source_props.created
            if source_props.identifier:
                target_props.identifier = source_props.identifier
            if source_props.keywords:
                target_props.keywords = source_props.keywords
            if source_props.language:
                target_props.language = source_props.language
            if source_props.last_modified_by:
                target_props.last_modified_by = source_props.last_modified_by
            if source_props.last_printed:
                target_props.last_printed = source_props.last_printed
            if source_props.modified:
                target_props.modified = source_props.modified
            if source_props.revision:
                target_props.revision = source_props.revision
            if source_props.subject:
                target_props.subject = source_props.subject
            if source_props.title:
                target_props.title = source_props.title
            if source_props.version:
                target_props.version = source_props.version
            
            # Зберігаємо презентацію
            target_prs.save(target_path)
            logging.debug(f"PPTX metadata copied for {target_path}")
            
        except Exception as e:
            logging.debug(f"Could not copy PPTX metadata using python-pptx: {e}")
            # Fallback до копіювання XML
            self._copy_office_metadata_xml(source_path, target_path)
    
    def _copy_image_metadata(self, source_path, target_path):
        """
        Копіює EXIF метадані зображення
        
        Args:
            source_path: Шлях до оригінального файлу
            target_path: Шлях до цільового файлу
        """
        try:
            from PIL import Image
            
            # Відкриваємо обидва зображення
            source_img = Image.open(source_path)
            target_img = Image.open(target_path)
            
            # Копіюємо EXIF дані
            if hasattr(source_img, 'info') and source_img.info:
                # Зберігаємо з метаданими оригіналу
                target_img.save(target_path, exif=source_img.info.get('exif', b''))
            
            logging.debug(f"Image metadata copied for {target_path}")
            
        except Exception as e:
            logging.debug(f"Could not copy image metadata: {e}")
    
    # --- ОБРОБКА ФАЙЛІВ ---
    
    def start_corruption_thread(self):
        """Запуск процесу в окремому потоці (corruption або secure delete)"""
        if not self.selected_files:
            messagebox.showwarning("Warning", "No files added!")
            return
        
        self.btn_run.configure(state="disabled", text="Processing...")
        self.progress_bar.set(0)
        
        operation = self.operation_mode.get()
        logging.info(f"Starting {operation} for {len(self.selected_files)} files")
        
        threading.Thread(target=self.process_files, daemon=True).start()
    
    def process_files(self):
        """Обробка файлів (виконується в окремому потоці)"""
        operation = self.operation_mode.get()
        
        if operation == OperationMode.CORRUPT:
            self.process_corruption()
        elif operation == OperationMode.SECURE_DELETE:
            self.process_secure_delete()
    
    def process_corruption(self):
        """Обробка corruption режиму"""
        total = len(self.selected_files)
        step = 1 / total if total > 0 else 0
        current_progress = 0
        
        success_count = 0
        error_count = 0
        
        # Скидання статистики
        self.total_stats = {
            'total_files': 0,
            'total_bytes_changed': 0,
            'total_time': 0.0,
            'original_total_size': 0,
            'corrupted_total_size': 0
        }
        
        # Створення corruption engine
        custom_params = {
            'header_bytes': self.custom_header_bytes.get(),
            'noise_count': self.custom_noise_count.get(),
            'truncate_bytes': self.custom_truncate_bytes.get()
        }
        engine = CorruptionEngine(self.corruption_mode.get(), custom_params)
        
        for file_path in self.selected_files:
            try:
                # Визначення шляху для нового файлу
                if self.output_directory.get():
                    directory = self.output_directory.get()
                else:
                    directory = os.path.dirname(file_path)
                
                # Генерація імені за шаблоном
                new_path = self.generate_output_filename(file_path, directory)
                
                # Оновлення статусу
                self.after(0, self.update_file_status, file_path, "Processing...", 
                          AppConfig.COLOR_PROCESSING)
                
                # Копіювання БЕЗ метаданих (просто байти)
                with open(file_path, 'rb') as src:
                    content = src.read()
                with open(new_path, 'wb') as dst:
                    dst.write(content)
                
                logging.info(f"Created copy: {new_path}")
                
                # Застосування пошкодження та збір статистики
                stats = engine.corrupt_file(new_path)
                
                # ПІСЛЯ corruption відновлюємо метадані з оригіналу
                self.preserve_metadata(file_path, new_path)
                
                # Оновлення загальної статистики
                self.total_stats['total_files'] += 1
                self.total_stats['total_bytes_changed'] += stats.bytes_changed
                self.total_stats['total_time'] += stats.processing_time
                self.total_stats['original_total_size'] += stats.original_size
                self.total_stats['corrupted_total_size'] += stats.corrupted_size
                
                success_count += 1
                
                # Детальна інформація у статусі
                status_text = f"✓ {stats.processing_time:.2f}s"
                self.after(0, self.update_file_status, file_path, status_text, 
                          AppConfig.COLOR_SUCCESS)
                
                logging.info(f"Successfully corrupted: {new_path} | "
                           f"Changed: {stats.bytes_changed}B | Time: {stats.processing_time:.3f}s")
            
            except PermissionError as e:
                error_count += 1
                self.after(0, self.update_file_status, file_path, "Access Denied", 
                          AppConfig.COLOR_ERROR)
                logging.error(f"Permission denied: {file_path} - {e}")
            
            except IOError as e:
                error_count += 1
                self.after(0, self.update_file_status, file_path, "I/O Error", 
                          AppConfig.COLOR_ERROR)
                logging.error(f"I/O error: {file_path} - {e}")
            
            except Exception as e:
                error_count += 1
                self.after(0, self.update_file_status, file_path, "Error", 
                          AppConfig.COLOR_WARNING)
                logging.error(f"Unexpected error processing {file_path}: {e}", exc_info=True)
            
            current_progress += step
            self.after(0, self.progress_bar.set, current_progress)
        
        # Завершення
        self.after(0, self.on_processing_complete, success_count, error_count)
    
    def process_secure_delete(self):
        """Обробка secure delete режиму"""
        total = len(self.selected_files)
        step = 1 / total if total > 0 else 0
        current_progress = 0
        
        success_count = 0
        error_count = 0
        
        # Скидання статистики
        self.total_stats = {
            'total_files': 0,
            'total_bytes_changed': 0,
            'total_time': 0.0,
            'original_total_size': 0,
            'corrupted_total_size': 0
        }
        
        # Створення SecureShredder з параметрами користувача
        passes = self.deletion_passes.get()
        shredder = SecureShredder(passes=passes)
        
        logging.info(f"Starting secure deletion with {passes} passes")
        
        for file_path in self.selected_files:
            try:
                # Оновлення статусу
                self.after(0, self.update_file_status, file_path, "Shredding...", 
                          AppConfig.COLOR_PROCESSING)
                
                # Запуск безпечного видалення
                stats = shredder.shred_file(file_path)
                
                # Оновлення загальної статистики
                if stats.success:
                    self.total_stats['total_files'] += 1
                    self.total_stats['total_time'] += stats.processing_time
                    self.total_stats['original_total_size'] += stats.file_size
                    
                    success_count += 1
                    
                    # Детальна інформація у статусі
                    status_text = f"✓ {stats.processing_time:.2f}s"
                    self.after(0, self.update_file_status, file_path, status_text, 
                              AppConfig.COLOR_SUCCESS)
                    
                    logging.info(f"Successfully shredded: {file_path} | "
                               f"Size: {self.format_file_size(stats.file_size)} | "
                               f"Time: {stats.processing_time:.3f}s")
                else:
                    error_count += 1
                    self.after(0, self.update_file_status, file_path, "Failed", 
                              AppConfig.COLOR_ERROR)
                    logging.error(f"Failed to shred: {file_path}")
            
            except PermissionError as e:
                error_count += 1
                self.after(0, self.update_file_status, file_path, "Access Denied", 
                          AppConfig.COLOR_ERROR)
                logging.error(f"Permission denied: {file_path} - {e}")
            
            except IOError as e:
                error_count += 1
                self.after(0, self.update_file_status, file_path, "I/O Error", 
                          AppConfig.COLOR_ERROR)
                logging.error(f"I/O error: {file_path} - {e}")
            
            except Exception as e:
                error_count += 1
                self.after(0, self.update_file_status, file_path, "Error", 
                          AppConfig.COLOR_WARNING)
                logging.error(f"Unexpected error: {file_path}: {e}", exc_info=True)
            
            current_progress += step
            self.after(0, self.progress_bar.set, current_progress)
        
        # Завершення
        self.after(0, self.on_processing_complete_delete, success_count, error_count)
    
    def update_file_status(self, file_path, status_text, color):
        """Оновлення статусу файлу"""
        if file_path in self.file_rows:
            self.file_rows[file_path].status_label.configure(text=status_text, text_color=color)
    
    def on_processing_complete(self, success_count, error_count):
        """Завершення обробки (corruption)"""
        self.btn_run.configure(state="normal", text="Corrupt All →")
        
        # Формування детального звіту
        stats = self.total_stats
        avg_time = stats['total_time'] / stats['total_files'] if stats['total_files'] > 0 else 0
        size_diff = stats['original_total_size'] - stats['corrupted_total_size']
        
        message = f"Processing complete!\n\n"
        message += f"✓ Success: {success_count}\n"
        message += f"✗ Errors: {error_count}\n\n"
        message += f"📊 Statistics:\n"
        message += f"• Total bytes changed: {self.format_file_size(stats['total_bytes_changed'])}\n"
        message += f"• Original size: {self.format_file_size(stats['original_total_size'])}\n"
        message += f"• Corrupted size: {self.format_file_size(stats['corrupted_total_size'])}\n"
        message += f"• Size reduced: {self.format_file_size(size_diff)}\n"
        message += f"• Total time: {stats['total_time']:.2f}s\n"
        message += f"• Avg per file: {avg_time:.3f}s"
        
        if error_count == 0:
            message += "\n\n✅ All files processed successfully!"
        
        logging.info(f"Processing complete: {success_count} success, {error_count} errors")
        logging.info(f"Statistics: {stats['total_bytes_changed']} bytes changed, "
                    f"{stats['total_time']:.2f}s total, {avg_time:.3f}s avg")
        
        messagebox.showinfo("Done", message)
    
    def on_processing_complete_delete(self, success_count, error_count):
        """Завершення обробки (secure delete)"""
        self.btn_run.configure(state="normal", text="Secure Delete →")
        
        # Формування детального звіту
        stats = self.total_stats
        avg_time = stats['total_time'] / stats['total_files'] if stats['total_files'] > 0 else 0
        
        message = f"Secure deletion complete!\n\n"
        message += f"✓ Success: {success_count}\n"
        message += f"✗ Errors: {error_count}\n\n"
        message += f"📊 Statistics:\n"
        message += f"• Total files deleted: {stats['total_files']}\n"
        message += f"• Total data wiped: {self.format_file_size(stats['original_total_size'])}\n"
        message += f"• Total time: {stats['total_time']:.2f}s\n"
        message += f"• Avg per file: {avg_time:.3f}s"
        
        if error_count == 0:
            message += "\n\n✅ All files securely deleted!"
        
        logging.info(f"Secure deletion complete: {success_count} success, {error_count} errors")
        logging.info(f"Statistics: {self.format_file_size(stats['original_total_size'])} deleted, "
                    f"{stats['total_time']:.2f}s total")
        
        messagebox.showinfo("Done", message)
        
        # Формування детального звіту
        stats = self.total_stats
        avg_time = stats['total_time'] / stats['total_files'] if stats['total_files'] > 0 else 0
        size_diff = stats['original_total_size'] - stats['corrupted_total_size']
        
        message = f"Processing complete!\n\n"
        message += f"✓ Success: {success_count}\n"
        message += f"✗ Errors: {error_count}\n\n"
        message += f"📊 Statistics:\n"
        message += f"• Total bytes changed: {self.format_file_size(stats['total_bytes_changed'])}\n"
        message += f"• Original size: {self.format_file_size(stats['original_total_size'])}\n"
        message += f"• Corrupted size: {self.format_file_size(stats['corrupted_total_size'])}\n"
        message += f"• Size reduced: {self.format_file_size(size_diff)}\n"
        message += f"• Total time: {stats['total_time']:.2f}s\n"
        message += f"• Avg per file: {avg_time:.3f}s"
        
        if error_count == 0:
            message += "\n\n✅ All files processed successfully!"
        
        logging.info(f"Processing complete: {success_count} success, {error_count} errors")
        logging.info(f"Statistics: {stats['total_bytes_changed']} bytes changed, "
                    f"{stats['total_time']:.2f}s total, {avg_time:.3f}s avg")
        
        messagebox.showinfo("Done", message)
